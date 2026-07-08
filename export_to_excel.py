"""
export_to_excel.py
==================
Đọc các file JSON hóa đơn tiền điện từ thư mục output_bill,
rồi xuất ra file Excel với mỗi sheet tương ứng một khách hàng (Mã KH).

Nội dung mỗi sheet khách hàng:
    - Metadata khách hàng (Mã KH, Tên KH, Địa chỉ) ở đầu trang.
    - Bảng chi tiết gồm các cột:
        - Kỳ hóa đơn
        - Tổng kWh
        - Tổng tiền thanh toán (VNĐ)
    - Hàng Tổng cộng (TỔNG CỘNG) tính tổng kWh và tổng tiền của các kỳ.

Yêu cầu:
    pip install openpyxl

Sử dụng:
    python export_to_excel.py
    python export_to_excel.py --input-dir output_bill --output hoa_don.xlsx
"""

import argparse
import glob
import json
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] Thư viện 'openpyxl' chưa được cài đặt.")
    print("        Chạy:  pip install openpyxl")
    sys.exit(1)


# =========================
# LOGGING
# =========================
def log(message: str) -> None:
    try:
        print(message, flush=True)
    except UnicodeEncodeError:
        sys.stdout.buffer.write((message + "\n").encode("utf-8", errors="replace"))
        sys.stdout.buffer.flush()


# =========================
# CONFIG
# =========================
DEFAULT_INPUT_DIR = os.path.join(os.path.dirname(__file__), "ah_raw_bill")
DEFAULT_OUTPUT = os.path.join(os.path.dirname(__file__), "hoa_don_export.xlsx")

# Màu header (xanh dương đậm)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# Màu tổng cộng
TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FONT = Font(bold=True, size=11)

# Viền ô
THIN_SIDE = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# =========================
# HELPERS
# =========================
def clean_html(text: str) -> str:
    """Loại bỏ các thẻ HTML như </div> khỏi chuỗi."""
    if not text:
        return ""
    return re.sub(r"<[^>]+>", "", text).strip()


def parse_end_date(ky_hoa_don: str) -> str:
    """
    Trích ngày kết thúc kỳ từ chuỗi ky_hoa_don.
    Ví dụ: "Kỳ 1 - 6/2026 (10 ngày từ 01/06/2026 đến 10/06/2026)" -> "10/06/2026"
    """
    cleaned = clean_html(ky_hoa_don)
    dates = re.findall(r"\d{2}/\d{2}/\d{4}", cleaned)
    return dates[-1] if dates else ""


def sort_key_for_date(date_str: str) -> datetime:
    """Chuyển dd/mm/yyyy thành datetime để sắp xếp."""
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        return datetime.min


def sort_key_for_client_record(client: dict) -> datetime:
    """Sắp xếp các kỳ của khách hàng tăng dần theo ngày kết thúc."""
    ky = client.get("ky_hoa_don", "")
    end_date = parse_end_date(ky)
    return sort_key_for_date(end_date)


# =========================
# EXCEL STYLING HELPERS
# =========================
def apply_header(ws, headers: list[str], row: int = 5):
    """Ghi và tô màu hàng header bảng."""
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def apply_data_row(ws, row: int, values: list, alignments: list):
    """Ghi một hàng dữ liệu với căn lề và viền."""
    for col_idx, (val, align) in enumerate(zip(values, alignments), start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = align
        cell.border = THIN_BORDER


def apply_total_row(ws, row: int, values: list, alignments: list):
    """Ghi hàng tổng cộng với màu nền khác."""
    for col_idx, (val, align) in enumerate(zip(values, alignments), start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.alignment = align
        cell.border = THIN_BORDER


def auto_fit_columns(ws):
    """Tự động điều chỉnh độ rộng cột theo nội dung."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                # Nếu là dòng 1-3 thông tin khách hàng, không dùng để tính độ rộng cột 2 
                # để tránh cột quá rộng do địa chỉ dài.
                if row in (1, 2, 3) and col_idx == 2:
                    continue
                max_len = max(max_len, len(str(val)))
        
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 15)


# =========================
# CORE LOGIC
# =========================
def load_all_json(input_dir: str) -> list[dict]:
    """Đọc tất cả file JSON trong input_dir và gộp danh sách clients."""
    pattern = os.path.join(input_dir, "*.json")
    json_files = sorted(glob.glob(pattern))
    if not json_files:
        log(f"[ERROR] Không tìm thấy file JSON nào trong: {input_dir}")
        sys.exit(1)

    log(f"[INFO] Tìm thấy {len(json_files)} file JSON:")
    all_clients: list[dict] = []
    for path in json_files:
        log(f"       - {os.path.basename(path)}")
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        clients = data.get("clients", [])
        all_clients.extend(clients)

    log(f"[INFO] Tổng số bản ghi tìm thấy: {len(all_clients)}")
    return all_clients


def group_by_client(clients: list[dict]) -> dict[str, list[dict]]:
    """
    Nhóm clients theo mã khách hàng.
    Key: mã khách hàng (vd: "PE15000352029")
    """
    groups: dict[str, list[dict]] = {}
    for client in clients:
        if "error" in client:
            continue  # bỏ qua bản ghi lỗi
        ma_kh = client.get("ma_khach_hang", "").strip()
        if not ma_kh:
            continue
        groups.setdefault(ma_kh, []).append(client)
    return groups


def build_excel(groups: dict[str, list[dict]], output_path: str):
    """Tạo file Excel với mỗi khách hàng một sheet."""
    wb = openpyxl.Workbook()
    # Xóa sheet mặc định
    wb.remove(wb.active)

    # Sắp xếp mã khách hàng theo bảng chữ cái
    sorted_client_codes = sorted(groups.keys())

    HEADERS = ["Kỳ hóa đơn", "Tổng kWh", "Tổng tiền thanh toán (VNĐ)"]
    ALIGNMENTS = [LEFT, RIGHT, RIGHT]
    TOTAL_ALIGNMENTS = [CENTER, RIGHT, RIGHT]

    for ma_kh in sorted_client_codes:
        records = groups[ma_kh]
        
        # Sắp xếp các kỳ của khách hàng theo thời gian tăng dần
        records_sorted = sorted(records, key=sort_key_for_client_record)

        # Lấy thông tin khách hàng từ bản ghi đầu tiên
        first_rec = records_sorted[0]
        ten_kh = first_rec.get("ten_khach_hang", "")
        # Nếu trích xuất nhầm "Địa chỉ" làm tên, giữ nguyên hoặc bỏ qua
        if ten_kh.strip() == "Địa chỉ":
            ten_kh = "Chưa rõ"
        dia_chi = first_rec.get("dia_chi", "")

        # Tạo sheet mới cho khách hàng (Excel giới hạn tên sheet tối đa 31 ký tự)
        sheet_name = ma_kh[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 1. Ghi Metadata khách hàng ở các hàng đầu tiên
        ws.cell(row=1, column=1, value="Mã khách hàng:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=ma_kh)
        
        ws.cell(row=2, column=1, value="Tên khách hàng:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=ten_kh)
        
        ws.cell(row=3, column=1, value="Địa chỉ:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=dia_chi)

        # Căn lề trái cho metadata
        for r in (1, 2, 3):
            ws.cell(row=r, column=1).alignment = LEFT
            ws.cell(row=r, column=2).alignment = LEFT

        # Cố định hàng header (hàng 5)
        ws.freeze_panes = "A6"

        # 2. Ghi Header bảng ở hàng 5
        apply_header(ws, HEADERS, row=5)

        # 3. Ghi dữ liệu các kỳ hóa đơn
        total_kwh = 0
        total_tien = 0
        start_row = 6

        for i, client in enumerate(records_sorted):
            row_idx = start_row + i
            ky_raw = client.get("ky_hoa_don", "")
            ky_clean = clean_html(ky_raw)
            kwh = client.get("tong_kwh") or 0
            tien = client.get("tong_tien_thanh_toan")

            kwh_val = int(kwh)
            tien_val = int(tien) if tien is not None else None

            apply_data_row(
                ws,
                row=row_idx,
                values=[ky_clean, kwh_val, tien_val],
                alignments=ALIGNMENTS,
            )

            total_kwh += kwh_val
            if tien_val is not None:
                total_tien += tien_val

        # 4. Ghi hàng Tổng cộng
        total_row = start_row + len(records_sorted)
        apply_total_row(
            ws,
            row=total_row,
            values=["TỔNG CỘNG", total_kwh, total_tien],
            alignments=TOTAL_ALIGNMENTS,
        )

        # Định dạng số cho cột kWh (B) và tiền (C) cho các hàng dữ liệu + tổng cộng
        for row_idx in range(start_row, total_row + 1):
            ws.cell(row=row_idx, column=2).number_format = "#,##0"
            ws.cell(row=row_idx, column=3).number_format = "#,##0"

        # Thiết lập màu tab cho chuyên nghiệp
        ws.sheet_properties.tabColor = "1F4E79"

        # Auto-fit độ rộng cột
        auto_fit_columns(ws)

        log(f"   [OK] Khách hàng '{ma_kh}': {len(records_sorted)} kỳ hóa đơn")

    wb.save(output_path)
    log(f"\n[OK] Đã xuất file Excel thành công: {output_path}")


# =========================
# MAIN
# =========================
def main():
    parser = argparse.ArgumentParser(
        description="Xuất dữ liệu hóa đơn điện từ JSON ra Excel (mỗi khách hàng một sheet)"
    )
    parser.add_argument(
        "--input-dir",
        default=DEFAULT_INPUT_DIR,
        help=f"Thư mục chứa các file JSON đầu vào (default: {DEFAULT_INPUT_DIR})",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Đường dẫn file Excel đầu ra (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    log("=== Xuất hóa đơn điện -> Excel (Một sheet/khách hàng) ===")
    log(f"[INFO] Thư mục JSON : {args.input_dir}")
    log(f"[INFO] File Excel   : {args.output}")
    log("")

    clients = load_all_json(args.input_dir)
    groups = group_by_client(clients)

    log(f"\n[INFO] Số khách hàng: {len(groups)}")
    build_excel(groups, args.output)
    log("=== Hoàn thành ===")


if __name__ == "__main__":
    main()
